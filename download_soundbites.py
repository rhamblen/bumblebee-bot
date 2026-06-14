"""
Download reference audio clips from 101soundboards.com boards.
Saves to \\SERVER-UR1\media\bumblebee\references\{character_folder}\
Updates the spreadsheet with local file paths in a new column.
"""

import re
import os
import time
import urllib.request
import urllib.error
import json
import sys
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

BASE_URL   = "https://www.101soundboards.com"
DEST_ROOT  = Path(r"\\SERVER-UR1\media\bumblebee\references")
XLSX_PATH  = Path(r"D:\backup\richard\Documents\bumblebee bot\list of character voices.xlsx")
HEADERS    = {"User-Agent": "Mozilla/5.0", "Accept": "application/json"}

def sanitize(name: str) -> str:
    return re.sub(r'[\\/:*?"<>|]', '', name).strip().replace(' ', '_')[:60]

def fetch_json(url: str) -> dict:
    req = urllib.request.Request(url, headers=HEADERS)
    with urllib.request.urlopen(req, timeout=15) as r:
        return json.loads(r.read().decode())

def download_file(url: str, dest: Path) -> bool:
    if dest.exists():
        print(f"    already exists, skipping")
        return True
    try:
        req = urllib.request.Request(url, headers=HEADERS)
        with urllib.request.urlopen(req, timeout=30) as r, open(dest, 'wb') as f:
            f.write(r.read())
        return True
    except urllib.error.HTTPError as e:
        print(f"    HTTP {e.code} downloading {url}")
        return False
    except Exception as e:
        print(f"    Error: {e}")
        return False

def get_board_id(url: str):
    m = re.search(r'/boards/(\d+)', url)
    return int(m.group(1)) if m else None

# ── Read spreadsheet ──────────────────────────────────────────────────────────
wb = load_workbook(XLSX_PATH)
ws = wb.active

# Find column G (101Soundboards Link) — index 7 (1-based)
SB_COL     = 7   # G = col 7
PERSON_COL = 3   # C = Person name
ARCH_COL   = 1   # A = Voice Archetype

# Add new column H for local file paths (insert before current col H = TTS Prompt)
ws.insert_cols(8)
h_cell = ws.cell(row=1, column=8)
h_cell.value = "Local Reference Files"
h_cell.font  = Font(bold=True)
h_cell.fill  = PatternFill('solid', start_color='FFF2CC')
h_cell.alignment = Alignment(wrap_text=True)
ws.column_dimensions['H'].width = 60

results = {}  # row -> list of local paths

# ── Process each board row ────────────────────────────────────────────────────
for row in range(2, ws.max_row + 1):
    sb_url   = ws.cell(row=row, column=SB_COL).value or ''
    archetype = ws.cell(row=row, column=ARCH_COL).value or ''
    person    = ws.cell(row=row, column=PERSON_COL).value or ''

    if not sb_url.startswith('http') or '101soundboards' not in sb_url:
        continue

    if '/tts/' in sb_url:
        print(f"Row {row:3d} [{archetype}]: TTS board — skipping (AI-generated, no pre-recorded clips)")
        ws.cell(row=row, column=8).value = "TTS board — no downloadable clips"
        continue

    board_id = get_board_id(sb_url)
    if not board_id:
        print(f"Row {row:3d} [{archetype}]: could not parse board ID from {sb_url}")
        continue

    # Folder name: prefer person name, fall back to archetype
    folder_name = sanitize(person if person and person != 'nan' else archetype)
    dest_dir = DEST_ROOT / folder_name
    dest_dir.mkdir(parents=True, exist_ok=True)

    print(f"\nRow {row:3d} [{archetype}] -> board {board_id} -> {dest_dir}")

    # Fetch board sounds from API
    try:
        data = fetch_json(f"{BASE_URL}/api/v1/boards/{board_id}")
        sounds = data.get('data', {}).get('sounds', [])
    except Exception as e:
        print(f"    API error: {e}")
        ws.cell(row=row, column=8).value = f"API error: {e}"
        continue

    if not sounds:
        print(f"    No sounds found")
        ws.cell(row=row, column=8).value = "No sounds found"
        continue

    print(f"    {len(sounds)} sound(s) found")
    downloaded = []

    for sound in sounds:
        raw_url = sound.get('sound_file_url', '')
        file_url = raw_url.split('?')[0]   # strip any signed query params
        if not file_url:
            continue

        filename  = Path(file_url).name          # e.g. 3395799-aiding-each-other.mp3
        dest_path = dest_dir / filename
        full_url  = BASE_URL + file_url

        print(f"    Downloading: {filename}")
        ok = download_file(full_url, dest_path)
        if ok:
            downloaded.append(str(dest_path))
        time.sleep(0.3)   # polite rate-limit

    if downloaded:
        ws.cell(row=row, column=8).value = "\n".join(downloaded)
        print(f"    Saved {len(downloaded)} file(s)")
    else:
        ws.cell(row=row, column=8).value = "Download failed"

# ── Save workbook ─────────────────────────────────────────────────────────────
wb.save(XLSX_PATH)
print(f"\n✓ Spreadsheet updated: {XLSX_PATH}")
