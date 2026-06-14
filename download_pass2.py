"""
Second-pass downloader: handles full CDN URLs, fills in missing boards,
and updates the spreadsheet with all local file paths.
"""
import re, os, time, json, urllib.request, urllib.error
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

BASE_URL  = "https://www.101soundboards.com"
DEST_ROOT = Path(r"\\SERVER-UR1\media\bumblebee\references")
XLSX_PATH = Path(r"D:\backup\richard\Documents\bumblebee bot\list of character voices.xlsx")
HEADERS   = {"User-Agent": "Mozilla/5.0", "Accept": "application/json"}

def sanitize(name):
    return re.sub(r'[\\/:*?"<>|]', '', str(name)).strip().replace(' ', '_')[:60]

def fetch_json(url):
    req = urllib.request.Request(url, headers=HEADERS)
    with urllib.request.urlopen(req, timeout=15) as r:
        return json.loads(r.read().decode())

def resolve_url(raw_url):
    """Handle both relative (/storage/...) and full (https://hoovers...) URLs."""
    raw = raw_url.split('?')[0]
    if raw.startswith('http'):
        return raw
    return BASE_URL + raw

def download_file(url, dest):
    if dest.exists():
        return True, "exists"
    try:
        req = urllib.request.Request(url, headers=HEADERS)
        with urllib.request.urlopen(req, timeout=30) as r, open(dest, 'wb') as f:
            f.write(r.read())
        return True, "ok"
    except Exception as e:
        return False, str(e)[:60]

def board_id_from_url(url):
    m = re.search(r'/boards/(\d+)', url)
    return int(m.group(1)) if m else None

def folder_for(archetype, person):
    name = person if person and str(person).strip() not in ('', 'nan', 'None') else archetype
    s = sanitize(name)
    return s if s else 'unknown'

# ── Load workbook ─────────────────────────────────────────────────────────────
wb = load_workbook(XLSX_PATH)
ws = wb.active

# Columns (1-based): A=1 Archetype, C=3 Person, G=7 SBLink, H=8 Local Files
SB_COL    = 7
LOCAL_COL = 8

# Ensure column H header exists
if not ws.cell(row=1, column=LOCAL_COL).value:
    c = ws.cell(row=1, column=LOCAL_COL)
    c.value = "Local Reference Files"
    c.font  = Font(bold=True)
    c.fill  = PatternFill('solid', start_color='FFF2CC')
    c.alignment = Alignment(wrap_text=True)
    ws.column_dimensions['H'].width = 60

# Track which board IDs we've already assigned to a folder
board_folder_map = {}  # board_id -> folder path

for row in range(2, ws.max_row + 1):
    sb_url    = ws.cell(row=row, column=SB_COL).value or ''
    archetype = ws.cell(row=row, column=1).value or ''
    person    = ws.cell(row=row, column=3).value or ''
    current   = ws.cell(row=row, column=LOCAL_COL).value or ''

    if not sb_url.startswith('http') or '101soundboards' not in sb_url:
        continue

    if '/tts/' in sb_url:
        if not current:
            ws.cell(row=row, column=LOCAL_COL).value = "TTS board — no pre-recorded clips"
        continue

    board_id = board_id_from_url(sb_url)
    if not board_id:
        continue

    # If already done and has content, skip (unless it was 0 files)
    if current and 'failed' not in current.lower() and 'error' not in current.lower() and current.count('.mp3') > 0:
        # Check if files actually exist on disk
        paths = [p for p in current.splitlines() if p.strip()]
        if paths and Path(paths[0]).exists():
            print(f"Row {row:3d} [{archetype}]: already done ({current.count(chr(10))+1} files), skipping")
            continue

    # Reuse folder if same board was already used
    if board_id in board_folder_map:
        dest_dir = board_folder_map[board_id]
    else:
        folder_name = folder_for(archetype, person)
        dest_dir = DEST_ROOT / folder_name
        dest_dir.mkdir(parents=True, exist_ok=True)
        board_folder_map[board_id] = dest_dir

    print(f"\nRow {row:3d} [{archetype}] -> board {board_id} -> {dest_dir.name}")

    try:
        data   = fetch_json(f"{BASE_URL}/api/v1/boards/{board_id}")
        sounds = data.get('data', {}).get('sounds', [])
    except Exception as e:
        print(f"    API error: {e}")
        ws.cell(row=row, column=LOCAL_COL).value = f"API error: {e}"
        continue

    if not sounds:
        print(f"    No sounds in API response")
        ws.cell(row=row, column=LOCAL_COL).value = "No sounds found"
        continue

    print(f"    {len(sounds)} sound(s)")
    downloaded = []

    for sound in sounds:
        raw = sound.get('sound_file_url', '')
        if not raw:
            continue
        url      = resolve_url(raw)
        filename = Path(url).name
        dest     = dest_dir / filename
        ok, msg  = download_file(url, dest)
        status   = "OK" if ok else "FAIL"
        print(f"    {status} {filename[:55]}  {'' if ok else msg}")
        if ok:
            downloaded.append(str(dest))
        time.sleep(0.25)

    ws.cell(row=row, column=LOCAL_COL).value = "\n".join(downloaded) if downloaded else "Download failed"
    print(f"    >> {len(downloaded)} file(s) saved")

# ── Rename NA folder to a proper name ────────────────────────────────────────
na_dir = DEST_ROOT / "NA"
if na_dir.exists():
    proper = DEST_ROOT / "1950s_Newsreel_Vintage_Radio"
    if not proper.exists():
        na_dir.rename(proper)
        print(f"\nRenamed NA -> {proper.name}")
    # Update any cells pointing to the old NA path
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row=row, column=LOCAL_COL).value or ''
        if 'references\\NA\\' in val:
            ws.cell(row=row, column=LOCAL_COL).value = val.replace('references\\NA\\', f'references\\{proper.name}\\')

wb.save(XLSX_PATH)
print(f"\nDone. Spreadsheet saved.")
